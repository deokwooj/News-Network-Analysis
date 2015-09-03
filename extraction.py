#-*- coding:utf-8 -*-
import traceback
from konlpy.tag import Kkma
#from konlpy.utils import pprint
 
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

from openpyxl import load_workbook

from na_config import *

import pprint

def excel_noun():

	def excel_write(row_val, column_val, data):
		new_sheet.cell(row = row_val, column = column_val, value="%s" %data)

	wb=load_workbook(REFERENCE_EXCEL)
	sheetList = wb.get_sheet_names()
	sheet = wb.get_sheet_by_name(sheetList[0])
	row_count = sheet.get_highest_row()
	
	new_sheet = wb.create_sheet(title=EXTRACTION_SHEET)
	
	for i in range(2, row_count):
		if sheet.row_dimensions[i].visible :
			pass
		else :
			excel_write(i,1,'')
			new_sheet.row_dimensions[i].hidden = True
			#new_sheet.row_dimensions[i].outlineLevel = 1
			continue
	
		noun_val = ""
		full_qua = ""

		cellValue_name = sheet.cell(row=i, column=1).value
		cellValue = sheet.cell(row=i, column=2).value
		cellValue_article_id = sheet.cell(row=i, column=3).value

		try :
			QUA = cellValue.count(u'\u201c')
		except :
                        traceback.print_exc()
			#continue

		if QUA != -1:
			if QUA == 1 :
				START_QUA = cellValue.find(u"\u201c") + 1 # position of first quatation mark
				CELL_VALUE_LEN = len(cellValue)

				cellValue_re = cellValue[START_QUA:CELL_VALUE_LEN]
				END_QUA = cellValue_re.find(u"\u201d") # position of last quatation mark

				cellValue_final = cellValue_re[0:END_QUA]
				#print str(i) + "  "+ cellValue_name + "  "  + cellValue_final

				kkma = Kkma()
				#pprint (kkma.nouns(cellValue_final))
				s = (kkma.nouns(cellValue_final))

				for j in range(0,len(s)):
					noun_val = noun_val + s[j].encode('utf-8') + ','

				print noun_val

				excel_write(i, 1, cellValue_name)
				excel_write(i, 2, cellValue_final)
				excel_write(i, 3, noun_val)
				excel_write(i, 4, cellValue_article_id)

			elif QUA == 0 :
				#print str(i) + " " + cellValue
				ANOTHER_QUA = cellValue.find("\"") + 1 # position of first quatation mark
				ANOTHER_QUA_LEN = len(cellValue)

				another_cellValue = cellValue[ANOTHER_QUA:ANOTHER_QUA_LEN]
				ANOTHER_END_QUA = another_cellValue.find("\"")

				another_cellValue_final = another_cellValue[0:ANOTHER_END_QUA]
				#print str(i) + "  " + cellValue_name + "  " + another_cellValue_final
				kkma = Kkma()
				#pprint (kkma.nouns(cellValue_final))
				s = (kkma.nouns(another_cellValue_final))

				for j in range(0,len(s)):
					noun_val = noun_val + s[j].encode('utf-8') + ','

				excel_write(i, 1, cellValue_name)
				excel_write(i, 2, another_cellValue_final)
				excel_write(i, 3, noun_val)
				excel_write(i, 4, cellValue_article_id)

			elif QUA > 1 :
				#print str(i) + " " + str(QUA)
				try :
					for q in range(0,QUA):
						arr = cellValue.split(u"\u201d")
						arr_start_qua = arr[q].find(u"\u201c") + 1
						arr_len = len(arr[q]) 

						arr_cellValue = arr[q][arr_start_qua:arr_len]

						full_qua = full_qua + arr_cellValue

						kkma = Kkma()
						s = (kkma.nouns(arr_cellValue))

						for j in range(0,len(s)):
							noun_val = noun_val + s[j].encode('utf-8') + ','
							#print str(i) + " " + arr_cellValue

						excel_write(i, 1, cellValue_name)
						excel_write(i, 2, full_qua)
						excel_write(i, 3, noun_val)
						excel_write(i, 4, cellValue_article_id)
				except : 
                        		traceback.print_exc()

	wb.save(REFERENCE_EXCEL)


if __name__=="__main__":
	excel_noun()



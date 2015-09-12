#-*- coding: utf-8 -*-
from __future__ import division # To forace float point division

#########################################################
# Authors
#########################################################
# bla. bla...
#
#########################################################

#########################################################
# Liscence
#########################################################
# bla bla..
#
#########################################################

#########################################################
# Program Decsription
#########################################################
# this is news source analyis program. 
# read articles and discover network structure of news sources based on quataions in artticles. 
# Two stages :
# Stg 1. Excel files -->  Python data structure --> store binary format in hard disk as *.bin
# Stg 2. Load binary files, *.bin files into memory and performs data analytics with the loaded bin files. 
#
# bla bla...
#########################################################

#########################################################
# News Article Excel file source 
#########################################################
#1. reference.xlsx ('분단'에 대한 자료 엑셀 파일)
#2. wholetable.xlsx (정보원 자료 엑셀 파일)
#3. table_define.xlsx : 정보원 정의
#########################################################
"""
        table_define.xlsx : 정보원 정의
        | infoSrc_ID                   | 정보원 ID |
        | name                         | 이름 |
        | orgName                      | 소속이름 |
        | type                         | 정보원 구분 |
        | position                     | 직함 |
        | etc_position                 | 기타 직함 정보 |
        | yearOfBirth                  | 생년 |
        | person_id(FK)                | 사전의 인물 ID |
        | code                         | 인물의 소속 분류 |
        | is_classified_paper_category | 신문 지면 정보에 의해 정보원의 분류되었는지 여부 |
        | INFOSRC_GLOBAL_ID            | 전기간
        | infosrc_id_whole    | 5 |
        | infosrc_id_day      | 2003/10/10_408 |
        | infosrc_name        | 김수행 |
        | infosrc_org         | 서울대 |
        | infosrc_type        | I |
        | infosrc_pos         | 교수 |
        | infosrc_code        | 13 |
        | infosrc_isclassified| \N |        에 걸친 UniqueID |
        
Each column is extracted from wholetable.xlsx (정보원 자료 엑셀 파일)
 """
 
""" 
type
| S | 익명 - 소속 없는 사람 |
| R | 익명 - 소속 있는 사람 |
| I | 실명 개인 - 소속 있음 |
| N | 무속속 실명 |
| O | 조직 |
| s | 성만 나와 있는 익명 |        
"""

# Deokwoo Jung 's update 23 Aug by jdw-2. 
# modules to be imported
import os
import sys, traceback
import numpy as np
from numpy.linalg import inv
from numpy.linalg import norm
import uuid
import pylab as pl
from scipy import signal
from scipy import stats
from scipy.interpolate import interp1d
import matplotlib.pyplot as plt
from multiprocessing import Pool
#from datetime import datetime
import datetime as dt
from dateutil import tz
import shlex, subprocess
import time
import itertools
import calendar
import random
from matplotlib.collections import LineCollection
import pprint
import warnings
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

from openpyxl import load_workbook

from na_extraction import *
import cPickle as pickle

# newly imported
from na_config import *
from na_const import * 
import na_tools as nt


# Article is stroed in harddisk or DB,... 
QuoLabel={'pol':1, 'eco':2, 'tec':2,'cul':3, 'ent':3,'soc':4,'int':5, 'spo':6, 'etc':7}

TotalNum_Quotations=1000
# class definition : news source. 

# Dicitonary Definition
# Using 'set' data structure to extract a set of elements in each column of excel file. 
# Org_Name={1:'외무부', 2:'신한국당':,3:'서울대':,...}
# Job_Title={1: '변호사', 2:'사장':,3:'교수':,...};
# Function to extract sets from excel file. 


# R : no name yes org,    I : yes name yes org,    N : yes name no org,    O : org,    s : only last name
# NewsSource Type = {1:S, 2:R, 3:I, 4:N ,5:O, 6:s}

'''
def get_excel_sets(excel_dict):
    # load excel_dict
    name_set={} # Source's name
    org_set={} # Organization affiliated. 
    pos_set={} # Position held in the organization. 
    src_set={} # explain what it is  ???
    isc_set={} # explain what it is  ???
    return org_set,pos_set,src_set,isc_set
'''
######################################
# table_define.xlsx
# id : 정보원 ID
# name : 정보원 이름
# org : 정보원 조직
# srctype : 정보원 구분 {S,R,I,N,O,s}, (e.g. S ~ 익명 - 소속 없는 사람)
# pos : 정보원 직위
# code : 정보원 소속 분류, {헌법재판소, 재판부: 111 }, {검찰 : 211} 
# classified : 신문 지면 정보에 의해 정보원이 분류되어있는지 여부
######################################
class NewsSource:
    def __init__(self):
        self.id = None # uuid 
        self.date=dt.datetime(1999,12,31) #  quotations data,1999년 12월 31일 23시.
        self.name = None # name_set
        self.org = None # org_set
        self.srctype=None #~ {S,R,I,N,O,s}, (e.g. S ~ 익명 - 소속 없는 사람)
        self.pos = None #  Position 
        self.code=None # organization code
        self.classified=None # isclassified 
    def whoami(self): # print the current information for news source object
        for key in self.__dict__.items():
            print key[0],': ', key[1]\

######################################
# reference xlsx file
# 1 Reference sheet
#   - INFOSRC_NAME : 정보원 이름976.911 kB
#   - STN_CONTENT : 인용문이 들어간 문장
#   - ART_ID : 기사 ID
# 2 extraction sheet : 인용문 분리 후 명사 분리하고 정리
#   - 이름
#   - 인용문
#   - 명사
#   - 기사 ID
######################################
# Art.ID (meta_data_id) :"01101001[-->매체정보].20130527[-->날짜]100000112[-->기사ID] 
class NewsQuotation:
    def __init__(self):
        self.quotation_key =None # 4 digit number 
        self.article_id =None # 9 digit number 
        self.media_id = None  # 8 digit number string 
        self.date=dt.datetime(1999,12,31) #  quotations data,1999년 12월 31일 23시.
        self.news_src=NewsSource() # create NewsSource object
        self.quotation =None  # position, need utcto be initionalized by kkd_functions. 
        self.nounvec =None # position, need utcto be initionalized by kkd_functions. 
    def whoami(self):
        for key in self.__dict__.items():
            if key[0]=='news_src':
                print key[0],'name : ', key[1].name
            else:
                print key[0],': ', key[1]


def get_all_Quo():
    all_Quo=[]
    total_quo=TotalNum_Quotations
    #excel_nouns = pickle.load(open("./file/nouns.p","rb"))
    excel_nouns = pickle.load(open(DICT_NOUNS,"rb"))
    
    for i in range(0, len(excel_nouns)) :
        temp_nq = NewsQuotation() # create an instance of news quotations
        # To be manually or automatically (preferred)...
        temp_nq.gth_label = QuoLabel['eco'] # this is example. 
        #temp_nq.quo_nouns=get_nouns(temp_nq.quo_unicode) # to be done...
        temp_nq.quo_nouns = excel_nouns[i] # to be done...
        temp_nq.quo_date = '2015/07/30'# date of quotations , defined by datetime. 
        temp_nq.quo_article = '1' # Article Label ...

        all_Quo.append(temp_nq)
        #print all_Quo[i].quo_nouns
        # Fill all members and details...
        return all_Quo
 
# Return diatance matrix of Quatations
def generate_Qdist (w_param):
    # using w_param (weight coefficients for various distance matrix for Quatations 
    # w_param is given by DM's excel table ...
    return D_q


def quo_network_analysis(D_q,ns_param):
    # discover the best network structure given ns_param
    # D_q: Distance matrix for quoatations. 
    # 1. Clustering using D_q

    # 2. Applying na_param and cutoff neighbor max number of neighbor. 
    # 3. generate ns_structure = n by n binaryt matrix. 
    return ns_structutre
    

if __name__ == "__main__":
    print " running news source analysis....."

    # Load bin files for news sources and quotations 
    print 'Loading DICT_NEWS_INFO ...'
    # dict_news_info contains all quotations
    dict_news_info=nt.loadObjectBinaryFast(DICT_NEWS_INFO)
    
    NewsQuoObjs=[]
    ########################################################
    # Print all news dictionary information    
    ########################################################
    fld_name=('Name','Quatation', 'Nouns','Code')
    for k,(key, val) in enumerate(dict_news_info.iteritems()):
        quo_temp=NewsQuotation()
        quo_temp.quotation_key=str(key)
        for i,(fld_, val_) in enumerate(zip(fld_name,list(val)[0])):
            print fld_
            if fld_=='Name':
                quo_temp.news_src.name=val_
            elif fld_=='Quatation':
                val_temp=val_
                quo_temp.quotation=val_
            elif fld_=='Nouns':
                quo_temp.nounvec=val_
            elif fld_=='Code':
                val_temp=str(val_)[1:]
                quo_temp.media_id=val_temp[:8]
                quo_temp.date=\
                dt.datetime(int(val_temp[9:13]),int(val_temp[13:15]),int(val_temp[15:17]),23)
                quo_temp.article_id=val_temp[17:]
            else:
                warnings.warn("fld name not found")
        NewsQuoObjs.append((k,quo_temp))
    
   
    print '********************************************************************'
    print 'Print News Quotation Objects- NewsQuoObjs '    
    for (qid, obj_) in NewsQuoObjs:
        print '==================================================='
        print 'Quotation ID : ', qid 
        print '==================================================='
        obj_.whoami()
    print '********************************************************************'

    wb=load_workbook(WHOLETABLE_EXCEL)
    ws = wb.get_sheet_by_name('wholetable')
    NewsSrcObjs=[]
    src_fld_name=\
    ['src_id','src_date','src_name','src_org','src_type','src_pos','src_code','src_clfd']
    for k,row in enumerate(ws.iter_rows(row_offset=1)):
        src_temp=NewsSource()
        print k
        for (fld_, row_) in zip(src_fld_name,row):
            if row_.value=='null':
                row_.value=None
            elif row_.value=='\N':
                row_.value='N'
            val_temp=row_.value
            if val_temp!=None:
                if fld_=='src_id':
                    src_temp.id =val_temp
                elif fld_=='src_date':
                    val_temp=str(val_temp)[:10]
                    src_temp.date =dt.datetime(int(val_temp[:4]),int(val_temp[5:7]),int(val_temp[8:10]))
                elif fld_=='src_name':
                    src_temp.name =val_temp
                elif fld_=='src_org':
                    src_temp.org =val_temp
                elif fld_=='src_type':
                    src_temp.srctype =val_temp
                elif fld_=='src_pos':
                    src_temp.pos =val_temp
                elif fld_=='src_code':
                    src_temp.code =val_temp
                elif fld_=='src_clfd':                            
                    src_temp.classified =val_temp
                else:
                    warnings.warn("fld name not found")
        print  '==========================================='
        if src_temp.name!=None:
            NewsSrcObjs.append((k,src_temp))


    print '********************************************************************'
    print 'Print News Source Objects- NewsSrcObjs '    
    for (sid, obj_) in NewsSrcObjs:
        print '==================================================='
        print 'News Source  ID : ', sid 
        print '==================================================='
        obj_.whoami()
    print '********************************************************************'

    np.array([obj_.name for (sid, obj_) in NewsSrcObjs])
    print 'Generate News Source  '
    # News Sources by s = {s_1 , · · · , s_m }

    
    # Print list dictionary for news source. 
    print '------------------------------------'    
    print ' List of names in news sources '
    print '------------------------------------'

    
    
    # constrcut news source array 
    #for key in src_name.keys(): 
    #    print src_name[key]
    #print MyPrettyPrinter().pprint(src_name)
    print '------------------------------------'


    print '------------------------------------'    
    print ' List of organizations set '
    print '------------------------------------'

    print '------------------------------------'

    print '------------------------------------'    
    print ' List of position set '
    print '------------------------------------'
    #print MyPrettyPrinter().pprint(src_pos_set)
    print '------------------------------------'

    print '------------------------------------'    
    print ' List of organizations in news sources '
    print '------------------------------------'
    #for key in src_org.keys():
    #    print src_org[key]
    #print MyPrettyPrinter().pprint(src_org)
    print '------------------------------------'

   
    print '------------------------------------'    
    print ' List of positions in news sources '
    print '------------------------------------'
    #for key in src_pos.keys(): 
    #    print src_pos[key]
    #print MyPrettyPrinter().pprint(src_pos)
    print '------------------------------------'

    print '------------------------------------'    
    print ' List of informer class in news sources '
    print '------------------------------------'
    #for key in src_pos.keys(): 
    #    print src_pos[key]


    #print MyPrettyPrinter().pprint(test)

    # News Article by a = {a_1 , · · · , a_l }
    


    # Quotations in articles  by q = {q_1 , · · · , a_n }

    # U_{lxm} ~ Association matrix between News Sources S and Articles A

    # V_{mxn} ~ Association matrix  between News Sources S and Quotations Q.
    
    # Z_{nxl}~ Association matrix  between Quotations − Articles.
    
    # Q_v = V*V' 
    
    # Q_z = Z*Z'
    
    # \hat{q}_i ~ Projection of q_i into n-dimensional Euclidian space E_n    
    
    # D_q ~ Distance of matrix for \hat{q}_i s
    
    # D^_q = w_d D_q + w_v* Q_v + w_z*Q_z

    # Load class list of Quatation object. 
    #all_quo=get_all_Quo()
    
    # TODO: Create U, V, Z matrix here, (sample is ok ) --> check for paper. 
    # note: must be done after you clean your code with complete comments , 
    
    #####################################################
    # Simulation test 

    # This part is for simulations
    #####################################################
    
    #S-A associatoin matrix 
    
    #U=np.matrix(np.ones((5,2)))
    #U[3:5,0]=0
    #U[1:3,1]=0
    #S=U*U.T
    #pprint.pprint(S)
    
    print "version 2.56"
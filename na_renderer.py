# -*- coding: utf-8 -*-
# na_renderer.py  

import os
import uuid
import datetime
import numpy as np 
import math
import networkx as nx 
import matplotlib
import matplotlib.pyplot
import na_config as conf
from openpyxl import load_workbook
from openpyxl.workbook import Workbook 
from openpyxl.drawing import Image
from openpyxl.styles import Font, Style, Fill, Color, colors, fills, PatternFill


__author__ = "Jung-uk Choi"
__copyright__ = "Copyright 2015, Deokwoo Jung, All rights reserved."
__credits__ = ["Jung-uk Choi"]
__license__ = "GPL"
__version__ = "0.1"
__maintainer__ = "Jung-uk Choi"
__email__ = "choijunguk@gmail.com"
__status__ = "Prototype"


TMP_DIR = os.path.join('/tmp', 'NLP')
if not os.path.exists(TMP_DIR):
	os.system('mkdir ' + TMP_DIR)


def temporaryImagePath():
	return os.path.join(TMP_DIR, str(uuid.uuid4()) + '.png')


class Context(object):
	def __init__(self):
		super(Context, self).__init__()
		self.__quotations = {}
		self.__labels = {}
		self.__connectivity = None

	@property
	def quotations(self):
		return sorted(self.__quotations.keys())

	def quotationText(self, quotationID):
		assert isinstance(quotationID, (int, long))
		return self.__quotations[quotationID]['text']

	def quotationDate(self, quotationID):
		assert isinstance(quotationID, (int, long))
		return self.__quotations[quotationID]['date']

	def quotationArticleID(self, quotationID):
		assert isinstance(quotationID, (int, long))
		return self.__quotations[quotationID]['articleID']

	def quotationLabel(self, quotationID):
		assert isinstance(quotationID, (int, long))
		return self.__quotations[quotationID]['label']

	@property 
	def labels(self):
		return sorted(self.__labels.keys())

	def isExamplar(self, quotationID):
		assert isinstance(quotationID, (int, long))
		return self.__quotations[quotationID]['is_examplar']

	def labelExamplar(self, labelID):
		assert isinstance(labelID, (int, long))
		return self.__labels[labelID]['examplar']

	def labelQuotations(self, labelID):
		assert isinstance(labelID, (int, long))
		return self.__labels[labelID]['quotations']

	def setQuotationText(self, quotationID, text):
		assert isinstance(quotationID, (int, long))
		assert isinstance(text, (str, unicode))

		try:
			obj = self.__quotations[quotationID]
		except KeyError:
			obj = self.__quotations[quotationID] = {'text':None, 'label':None, 'is_examplar':False, 'date': None}

		obj['text'] = text

	def setQuotationDate(self, quotationID, date):
		assert isinstance(quotationID, (int, long))
		assert isinstance(date, (str, unicode, datetime.datetime))

		try:
			obj = self.__quotations[quotationID]
		except KeyError:
			obj = self.__quotations[quotationID] = {'text':None, 'label':None, 'is_examplar':False, 'date': None, 'articleID': None}

		obj['date'] = date

	def setQuotationArticleID(self, quotationID, articleID):
		assert isinstance(quotationID, (int, long))
		assert isinstance(articleID, (str, unicode))

		try:
			obj = self.__quotations[quotationID]
		except KeyError:
			obj = self.__quotations[quotationID] = {'text':None, 'label':None, 'is_examplar':False, 'date': None, 'articleID': None}

		obj['articleID'] = articleID

	def setQuotationLabel(self, quotationID, labelID):
		assert isinstance(quotationID, (int, long))
		assert isinstance(labelID, (int, long))

		try:
			obj = self.__labels[labelID]
		except KeyError:
			obj = self.__labels[labelID] = {'examplar': None, 'quotations': []}

		self.__quotations[quotationID]['label'] = labelID
		obj['quotations'].append(quotationID)

	def setLabelExamplar(self, labelID, quotationID):
		assert isinstance(labelID, (int, long))
		assert isinstance(quotationID, (int, long))

		try:
			obj = self.__labels[labelID]
		except KeyError:
			obj = self.__labels[labelID] = {'examplar': None, 'quotations': []}

		obj['examplar'] = quotationID
		self.__quotations[quotationID]['is_examplar'] = True

	def setConnectivityMatrix(self, m):
		assert isinstance(m, np.matrix)
		assert m.shape[0] == m.shape[1]

		self.__connectivity = m

	def connectivityMatrixForStepN(self, N):
		assert isinstance(N, int)
		assert N > 0

		if self.__connectivity is not None:
			m = self.__connectivity[:]
			for n in range(2, N+1):
				m += self.__connectivity ** n 

			for r in range(m.shape[0]):
				for c in range(m.shape[1]):
					if m[r, c]:
						m[r, c] = 1
					else:
						m[r, c] = 0

			return m

	def labelIdForQuotationID(self, quotationID):
		assert isinstance(quotationID, (int, long))	
		return self.__quotations[quotationID]['label']

	def neighbors(self, quotationID, N):
		assert isinstance(quotationID, (int, long))	
		assert isinstance(N, (int, long))

		neighbors = []

		quotationIDs = sorted(self.quotations)
		l1 = self.labelIdForQuotationID(quotationID)
		m = self.connectivityMatrixForStepN(N)
		for j in range(len(quotationIDs)):
			if self.isExamplar(quotationIDs[j]):
				l2 = self.labelIdForQuotationID(quotationIDs[j])
				if l1 != l2:
					if m[l1, l2]:
						neighbors.append(quotationIDs[j])

		return neighbors

	def degrees(self, quotationID, N):
		return len(self.neighbors(quotationID, N))


class ConnectedDots(object):
	def __init__(self, context):
		assert isinstance(context, Context)
		super(ConnectedDots, self).__init__()
		self.__context = context

	def render(self, N, outputPath=None):
		assert isinstance(N, int)
		assert N > 0
		assert outputPath is None or isinstance(outputPath, (str, unicode))

		matplotlib.pyplot.clf()

		G = nx.Graph()
		labels = {}
		degrees = {}

		quotationIDs = sorted(self.__context.quotations)

		# add all quotations
		for qid in quotationIDs:
			G.add_node(qid)
			labels[qid] = qid
			degrees[qid] = 0

		# add label - quotations relationship
		for l in self.__context.labels:
			for q in self.__context.labelQuotations(l):
				G.add_edge(self.__context.labelExamplar(l), q, weight=1.0)

		# add connectivity
		m = self.__context.connectivityMatrixForStepN(N)
		for i in range(len(quotationIDs)):
			q1 = quotationIDs[i]
			l1 = self.__context.labelIdForQuotationID(q1)

			for j in range(len(quotationIDs)):
				q2 = quotationIDs[j]
				l2 = self.__context.labelIdForQuotationID(q2)
				if l1 != l2:
					if m[l1, l2]:
						if self.__context.isExamplar(q1) and self.__context.isExamplar(q2):
							G.add_edge(q1, q2, weight=0.1)
							degrees[q1] += 1
							degrees[q2] += 1

		#
		sizes = []
		for qid in quotationIDs:
			d = degrees[qid]
			#sizes.append(500 + d * 100)
			sizes.append(500 + 3 * 100)

		#
		colors = []
		max_depth = float(max(degrees.values()))
		for qid in quotationIDs:
			#k = degrees[qid] / max_depth
			if self.__context.isExamplar(qid):
				colors.append('red')
			else:
				colors.append('orange')

		try:
			pos=nx.graphviz_layout(G)

		except:
			pos=nx.spring_layout(G, iterations=20)

		#pos = nx.circular_layout(G)
		#pos=nx.spring_layout(G, iterations=100)

		imageSize = math.sqrt(len(pos)) * 200 
		matplotlib.pyplot.figure(figsize=(imageSize/100,imageSize/100)) # image size
		matplotlib.pyplot.axis('equal')

		# nodes
		nx.draw_networkx_nodes(G, pos, nodelist=quotationIDs, node_size=sizes, font_size=5, node_color=colors)

		# edges
		elarge=[(u,v) for (u,v,d) in G.edges(data=True) if 'weight' in d and d['weight'] > 0.5]
		esmall=[(u,v) for (u,v,d) in G.edges(data=True) if 'weight' in d and d['weight'] <=0.5]
		nx.draw_networkx_edges(G, pos, edgelist=elarge, width=0.5, edge_color='gray')
		nx.draw_networkx_edges(G, pos, edgelist=esmall, width=2.0, edge_color='gray', style='dashed')

		nx.draw_networkx_labels(G, pos, labels)
		if outputPath:
			matplotlib.pyplot.savefig(outputPath)
		
		else:
			matplotlib.pyplot.show()


class ExcelRenderer(object):
	def __init__(self, context):
		assert isinstance(context, Context)
		super(ExcelRenderer, self).__init__()
		self.__context = context
		self.__connectedDotsRenderer = ConnectedDots(context)
		self.__columnStyle = Style(fill=PatternFill(patternType=fills.FILL_SOLID, fgColor=colors.BLACK), font=Font(bold=True, color=colors.WHITE))


	def render(self, outputPath, maxN):
		assert isinstance(outputPath, (str, unicode))
		assert isinstance(maxN, int)
		
		wb=Workbook()
		wb.remove_sheet(wb.worksheets[0])
		self.renderQuotationSheet(wb)
		for i in range(maxN):
			self.renderConnectivityMatrixSheet(wb, i+1)
			self.renderConnectivityChartSheet(wb, i+1)

		wb.save(outputPath)

	def renderQuotationSheet(self, wb):
		ws = wb.create_sheet()
		ws.title = 'Quotations'

		N = conf.maxConnectivity()

		ws.cell(row=1, column=1).value = 'N = %d, sim_thr = %f, MAX_NUM_QUO_ROWS = %d' % (conf.maxConnectivity(), conf.similarityThreshold(), conf.maxNumberOfQuotationRows())
		for i in range(6 + N * 2):
			ws.cell(row=1, column=1 + i).style = Style(fill=PatternFill(patternType=fills.FILL_SOLID, fgColor=colors.YELLOW), font=Font(bold=True, color=colors.RED))

		ws.cell(row=2, column=1).value = 'Quotation Key'
		ws.cell(row=2, column=2).value = 'Text'
		ws.cell(row=2, column=3).value = 'Label'
		ws.cell(row=2, column=4).value = 'Exemplar'

		for i in range(1, N + 1):
			ws.cell(row=2, column=3 + i * 2).value = 'N=%d (degree)' % i
			ws.cell(row=2, column=4 + i * 2).value = 'N=%d (neighbors)' % i


		ws.cell(row=2, column=5 + N * 2).value = 'Date'
		ws.cell(row=2, column=6 + N * 2).value = 'ArticleID'

		for c in range(6 + N * 2):
			ws.cell(row=2, column=1 + c).style = self.__columnStyle

		quotations = self.__context.quotations
		for (idx, qid) in enumerate(quotations):
			text = self.__context.quotationText(qid)
			label = self.__context.quotationLabel(qid)
			is_examplar = self.__context.isExamplar(qid)

			ws.cell(row=idx+3, column=1).value = qid
			ws.cell(row=idx+3, column=2).value = text
			ws.cell(row=idx+3, column=3).value = label

			if is_examplar:
				ws.cell(row=idx+3, column=4).value = "YES"
			else:
				ws.cell(row=idx+3, column=4).value = "NO"

			for n in range(1, N + 1):
				neighbors = self.__context.neighbors(qid, n)
				ws.cell(row=idx+3, column=3 + n * 2).value = len(neighbors)
				ws.cell(row=idx+3, column=4 + n * 2).value = repr(tuple(neighbors))

			ws.cell(row=idx+3, column=5 + N * 2).value = self.__context.quotationDate(qid)
			ws.cell(row=idx+3, column=6 + N * 2).value = self.__context.quotationArticleID(qid)

		return ws

	def renderConnectivityMatrixSheet(self, wb, N):
		ws = wb.create_sheet()
		ws.title = 'C.Matrix(N=%d)' % N

		labelIDs = self.__context.labels
		cell = ws.cell(row=0 + 1, column=0 + 1)
		cell.style = self.__columnStyle
		for i in range(len(labelIDs)):
			cell = ws.cell(row = 0 + 1, column = i+1 + 1)
			cell.value = self.__context.labelExamplar(labelIDs[i])
			cell.style = self.__columnStyle

			cell = ws.cell(row = i+1 + 1, column = 0 + 1)
			cell.value = self.__context.labelExamplar(labelIDs[i])
			cell.style = self.__columnStyle

		m = self.__context.connectivityMatrixForStepN(N)
		for r in range(m.shape[0]):
			for c in range(m.shape[1]):
				ws.cell(row = r + 1 + 1, column = c + 1 + 1).value = m[r, c]

		return ws

	def renderConnectivityChartSheet(self, wb, N):
		ws = wb.create_sheet()
		ws.title = 'C.Chart(N=%d)' % N

		path = temporaryImagePath()
		self.__connectedDotsRenderer.render(N, path)
		
		# image
		img = Image(path)
		img.anchor(ws['A1'])
		ws.add_image(img)
		return ws



if __name__ == '__main__':
	# q_id={0:23, 1:10, 2:39, 3:44, 4:14, 5:33, 6:21, 7:66, 8:88, 9:11}
	# q_label={0:0, 1:4, 2:1, 3:2, 4:3, 5:4, 6:1, 7:2, 8:2, 9:1}
	# q_exemplar={0:23, 1:39, 2:44, 3:14, 4:33}

	G_q = np.matrix([[1, 0, 1, 0, 0],
		             [0, 1, 1, 0, 1],
	                 [1, 1, 1, 1, 0],
	                 [0, 0, 1, 1, 0],
	                 [0, 1, 0, 0, 1]])

	obj = Context()
	obj.setQuotationText(23, 'text23')
	obj.setQuotationText(10, 'text10')
	obj.setQuotationText(39, 'text39')
	obj.setQuotationText(44, 'text44')
	obj.setQuotationText(14, 'text14')
	obj.setQuotationText(33, 'text33')
	obj.setQuotationText(21, 'text21')
	obj.setQuotationText(66, 'text66')
	obj.setQuotationText(88, 'text88')
	obj.setQuotationText(11, 'text11')

	obj.setQuotationLabel(23, 0)
	obj.setQuotationLabel(10, 4)
	obj.setQuotationLabel(39, 1)
	obj.setQuotationLabel(44, 2)
	obj.setQuotationLabel(14, 3)
	obj.setQuotationLabel(33, 4)
	obj.setQuotationLabel(21, 1)
	obj.setQuotationLabel(66, 2)
	obj.setQuotationLabel(88, 2)
	obj.setQuotationLabel(11, 1)

	obj.setLabelExamplar(0, 23)
	obj.setLabelExamplar(1, 39)
	obj.setLabelExamplar(2, 44)
	obj.setLabelExamplar(3, 14)
	obj.setLabelExamplar(4, 33)

	obj.setConnectivityMatrix(G_q)

	r2 = ExcelRenderer(obj)
	r2.render(os.path.join(os.getcwd(), 'output/output.xlsx'), 4)

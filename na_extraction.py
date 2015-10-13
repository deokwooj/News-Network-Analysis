#-*- coding:utf-8 -*-
import traceback
from konlpy.tag import Kkma
#from konlpy.utils import pprint
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
from na_config import *
import pprint
import na_tools as nt


def excel_noun():

    def excel_write(row_val, column_val, data):
        new_sheet.cell(row = row_val, column = column_val, value="%s" %data)

    wb=load_workbook(REFERENCE_EXCEL)
    sheetList = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetList[0])
    row_count = sheet.get_highest_row()
    
    new_sheet = wb.create_sheet(title=EXTRACTION_SHEET)
    
    for i in range(2, row_count):
        if sheet.row_dimensions[i].visible :
            pass
        else :
            excel_write(i,1,'')
            new_sheet.row_dimensions[i].hidden = True
            #new_sheet.row_dimensions[i].outlineLevel = 1
            continue
    
        noun_val = ""
        full_qua = ""

        cellValue_name = sheet.cell(row=i, column=1).value
        cellValue = sheet.cell(row=i, column=2).value
        cellValue_article_id = sheet.cell(row=i, column=3).value

        try :
            QUA = cellValue.count(u'\u201c')
        except :
                        traceback.print_exc()
            #continue

        if QUA != -1:
            if QUA == 1 :
                START_QUA = cellValue.find(u"\u201c") + 1 # position of first quatation mark
                CELL_VALUE_LEN = len(cellValue)

                cellValue_re = cellValue[START_QUA:CELL_VALUE_LEN]
                END_QUA = cellValue_re.find(u"\u201d") # position of last quatation mark

                cellValue_final = cellValue_re[0:END_QUA]
                #print str(i) + "  "+ cellValue_name + "  "  + cellValue_final

                kkma = Kkma()
                #pprint (kkma.nouns(cellValue_final))
                s = (kkma.nouns(cellValue_final))

                for j in range(0,len(s)):
                    noun_val = noun_val + s[j].encode('utf-8') + ','

                print noun_val

                excel_write(i, 1, cellValue_name)
                excel_write(i, 2, cellValue_final)
                excel_write(i, 3, noun_val)
                excel_write(i, 4, cellValue_article_id)

            elif QUA == 0 :
                #print str(i) + " " + cellValue
                ANOTHER_QUA = cellValue.find("\"") + 1 # position of first quatation mark
                ANOTHER_QUA_LEN = len(cellValue)

                another_cellValue = cellValue[ANOTHER_QUA:ANOTHER_QUA_LEN]
                ANOTHER_END_QUA = another_cellValue.find("\"")

                another_cellValue_final = another_cellValue[0:ANOTHER_END_QUA]
                #print str(i) + "  " + cellValue_name + "  " + another_cellValue_final
                kkma = Kkma()
                #pprint (kkma.nouns(cellValue_final))
                s = (kkma.nouns(another_cellValue_final))

                for j in range(0,len(s)):
                    noun_val = noun_val + s[j].encode('utf-8') + ','

                excel_write(i, 1, cellValue_name)
                excel_write(i, 2, another_cellValue_final)
                excel_write(i, 3, noun_val)
                excel_write(i, 4, cellValue_article_id)

            elif QUA > 1 :
                #print str(i) + " " + str(QUA)
                try :
                    for q in range(0,QUA):
                        arr = cellValue.split(u"\u201d")
                        arr_start_qua = arr[q].find(u"\u201c") + 1
                        arr_len = len(arr[q]) 

                        arr_cellValue = arr[q][arr_start_qua:arr_len]

                        full_qua = full_qua + arr_cellValue

                        kkma = Kkma()
                        s = (kkma.nouns(arr_cellValue))

                        for j in range(0,len(s)):
                            noun_val = noun_val + s[j].encode('utf-8') + ','
                            #print str(i) + " " + arr_cellValue

                        excel_write(i, 1, cellValue_name)
                        excel_write(i, 2, full_qua)
                        excel_write(i, 3, noun_val)
                        excel_write(i, 4, cellValue_article_id)
                except : 
                                traceback.print_exc()

    wb.save(REFERENCE_EXCEL)

def load_wholetable_excel():
    #wb=load_workbook('./file/wholetable.xlsx')
    wb=load_workbook(WHOLETABLE_EXCEL)
    sheetList = wb.get_sheet_names()
    #sheet = wb.get_sheet_by_name('wholetable')
    sheet = wb.get_sheet_by_name(WHOLETABLE_SHEET)
    return sheet


def load_reference_excel():
    #wb=load_workbook('./file/wholetable.xlsx')
    wb=load_workbook(REFERENCE_EXCEL)
    sheetList = wb.get_sheet_names()
    #sheet = wb.get_sheet_by_name('wholetable')
    sheet = wb.get_sheet_by_name(EXTRACTION_SHEET)
    return sheet

def extraction_tuple():
    sheet = load_reference_excel()
    row_count = sheet.get_highest_row()

    news_info = {}

    for i in range(2,row_count):
        if sheet.row_dimensions[i].visible :
            pass
        else :
            #new_sheet.row_dimensions[i].hidden = True
            #new_sheet.row_dimensions[i].outlineLevel = 
            continue

        news_source = sheet.cell(row=i, column=1).value 
        news_article = sheet.cell(row=i, column=2).value 
        news_noun = sheet.cell(row=i, column=3).value 
        news_article_id = sheet.cell(row=i, column=4).value 
        
        news_tuple=(news_source, news_article, news_noun, news_article_id)
        news_info[i]={news_tuple}

    nt.saveObjectBinaryFast(news_info, DICT_NEWS_INFO )


def org_set_dict():
    sheet = load_wholetable_excel()
    row_count = sheet.get_highest_row()

    dict_org_set={}
    org_items = set()

    for i in range(2,row_count):
        org = sheet.cell(row=i, column=4).value   # organization
        org_items.add(org)

    org_items = list(org_items)

    for j in range(0, len(org_items)):
        dict_org_set[j] = org_items[j]    # dictionary   index : organization
        #print dict_org_set[j]
        #print_dictionary(dict_org_set)

    return dict_org_set

def pos_set_dict():

    sheet = load_wholetable_excel()
    row_count = sheet.get_highest_row()

    dict_pos_set={}
    pos_items = set()

    for i in range(2,row_count):
        pos = sheet.cell(row=i, column=6).value   # 
        pos_items.add(pos)

    pos_items = list(pos_items)

    for j in range(0, len(pos_items)):
        dict_pos_set[j] = pos_items[j]    # dictionary   index : position 
    return dict_pos_set


def get_excel_type(i_type):
    if i_type == 'S':
        re_type = 1
    elif i_type == 'R':
        re_type = 2 
    elif i_type == 'I':
        re_type = 3
    elif i_type == 'N':
        re_type = 4
    elif i_type == 'O':
        re_type = 5
    elif i_type == 's':
        re_type = 6
    return re_type 


def get_excel_code():
    sheet = excel_open()
    row_count = sheet.get_highest_row()

    dict_code={}
    count = 0

    for i in range(2,row_count):
        #cell_ns=NewsSource() # create an instance of news sources
        code_tmp = sheet.cell(row=i, column=7).value   # type
        dict_code[str(count)] = str(code_tmp) 

        #print code_tmp , dict_code[str(count)] 
        count = count + 1

    #print_dictionary(dict_classified)

    count = 0
    return dict_code

def get_excel_classified(classified_tmp):

    if classified_tmp == '\\N':
        re_classified = 0 
    elif classified_tmp == '\\Y':
        re_classified= 1 

    #print classified_tmp + " " + str(dict_classified[str(count)])
    #print_dictionary(dict_classified)

    return re_classified 

####################################################
# Extract news source information
####################################################
def get_excel_informers():
    src_org_set = nt.loadObjectBinaryFast(DICT_ORG_SET)
    src_pos_set = nt.loadObjectBinaryFast(DICT_POS_SET)

    inv_src_org_set = {v: k for k, v in src_org_set.items()}
    inv_src_pos_set = {a: b for b, a in src_pos_set.items()}

    sheet = load_wholetable_excel()
    row_count = sheet.get_highest_row() 

    dict_id_name = {}
    dict_org = {}
    dict_type = {}
    dict_pos = {}
    dict_code = {}
    dict_classified = {}

    for i in range(2,row_count):
        id = sheet.cell(row=i, column=1).value   # id
        name = sheet.cell(row=i, column=3).value   # name
        org = sheet.cell(row=i, column=4).value   # organization
        i_type = sheet.cell(row=i, column=5).value   # organization
        pos = sheet.cell(row=i, column=6).value   # position
        code = sheet.cell(row=i, column=7).value   # organization
        classified = sheet.cell(row=i, column=8).value   # organization
        print code
        dict_id_name[id] = name   # dictionary id : name
        dict_type[id] = get_excel_type(i_type)   # dictionary   id : type
        dict_code[id] = code   # dictionary   id : code
        dict_classified[id] = get_excel_classified(classified)   # dictionary   id : classified

        try:
            idx_org = inv_src_org_set.keys().index(org)
            dict_org[id] = idx_org
        except ValueError:
            idx_org = -1

        try:
            idx_pos = inv_src_pos_set.keys().index(pos)
            dict_pos[id] = idx_pos
        except ValueError:
            idx_pos = -1

    return dict_id_name, dict_org, dict_type, dict_pos, dict_code, dict_classified 


def informer_class_dict():
    all_ns = []
# TODO: src_name is defined outside this function, !! please correct it
    for i in range(0, len(src_name)):
        ns_ins = NewsSource() # create an instance of news sources
        #ns_ins.id = src_name.keys.index[i] 
        #ns_ins.id = src.org.values()[i] 
        ns_ins.id = i 
        ns_ins.org = src_org.values()[i] 
        ns_ins.code = src_code.values()[i] 
        ns_ins.pos = src_pos.values()[i]
        ns_ins.i_type = src_type.values()[i] 
        ns_ins.classified = src_classified.values()[i]
        all_ns.append(ns_ins)

    return all_ns


    
def get_all_NS():
    #all_ns = pickle.load(open(DICT_INFORMER,"rb"))
    #TODO: dont hardcode constant, replace 6 with constant variables or get from class functions.   
    
    U=np.matrix(np.ones((len(src_mat),6)))
    
    for i in range(0, len(src_mat)):
        #print all_ns[i].id, all_ns[i].org, all_ns[i].pos
        #TODO: it is bad to use a, b in this for loop. 
        a=i
        b=i+1
        
        U[a:b,0]=src_mat[i].id
        U[a:b,1]=src_mat[i].org
        U[a:b,2]=src_mat[i].i_type
        U[a:b,3]=src_mat[i].pos
        U[a:b,4]=src_mat[i].code
        U[a:b,5]=src_mat[i].classified
    return U


def matrix_U():
    #all_ns = pickle.load(open(DICT_INFORMER,"rb"))
    # 메트릭스 0으로 초기화 
    
    U=np.matrix(np.zeros((len(src_n_informer_set),len(src_article_id_set))))
    sheet = load_reference_excel()
    row_count = sheet.get_highest_row()

    # 정보원이 속한 기사 ID 리스트
    test=[]

    # 중복되지 않은 정보원을 기중으로 for 루프
    for i in range(0, len(src_n_informer_set)):
        informer_tmp = src_n_informer_set[i]    # 정보원 set 딕션어리에 1번째 정보원 
        a = i
        b = i+1
        # 엑셀 파일에 정보원 비교하여 딕션어리 정보원과 같으면 그때의 기사 ID를 얻음
        for j in range(2,row_count):
            cell_informer = sheet.cell(row=j, column=1).value   # 
            if cell_informer is None:
                continue
            if informer_tmp == cell_informer:
                cell_article_id = sheet.cell(row=j, column=4).value   # 
                test.append(cell_article_id)
        # 생성된 메트릭스에 정보원이 기사 ID에 있으면 1을 넣음
        for k in range(0, len(src_article_id_set)):
            article_id_tmp = src_article_id_set[k]
            for m in range(0,len(test)):
                if article_id_tmp == test[m]:
                    U[a:b, k] = 1

        test[:] = [] #리스트 초기화  
    dump_matrix_U(U)


class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'    


def dump_matrix_U(U):
    # dump matrix

    print '=' * 80
    print "matrix U dump"
    for r in range(len(src_n_informer_set)):
        s = ''
        for c in range(len(src_article_id_set)):
            if U[r,c]:
                s += bcolors.OKBLUE
                s += bcolors.BOLD
                s += '%d' % U[r,c]
                s += bcolors.ENDC
                
            else:
                s += bcolors.FAIL
                s += '%d' % U[r,c]
                s += bcolors.ENDC

        print s
    print '=' * 80
    return U

def make_nouns_set():
    noun_items = set()
    dict_nouns = {}

    for i in range(0, len(excel_nouns_dict)):
        if excel_nouns_dict[i] is None:
            continue
        else :
            #print excel_nouns_dict[i].split(",")
            split_val = excel_nouns_dict[i].split(",")
            for noun_val in range(0, len(split_val)):
                noun_items.add(split_val[noun_val]) 
    return noun_items

def dump_matrix_V(U,y,x):
    # dump matrix

    print '=' * 80
    print "matrix V dump"
    for r in range(len(y)):
        s = ''
        for c in range(len(x)):
            if U[r,c]:
                s += bcolors.OKBLUE
                s += bcolors.BOLD
                s += '%d' % U[r,c]
                s += bcolors.ENDC
            else:
                s += bcolors.FAIL
                s += '%d' % U[r,c]
                s += bcolors.ENDC
        print s
    print '=' * 80
    return U

def matrix_V():
    make_nouns_set_test = make_nouns_set()

    U=np.matrix(np.zeros((len(src_split_arr_nouns),len(make_nouns_set_test))))

    V_test=[]

def matrix_V():
    make_nouns_set_test = make_nouns_set()

    #print make_nouns_set_test
    U=np.matrix(np.zeros((len(src_split_arr_nouns),len(make_nouns_set_test))))

    V_test=[]

    #MyPrettyPrinter().pprint(src_split_arr_nouns[0])
    #print len(src_split_arr_nouns)

    for i in range(0, len(src_split_arr_nouns)):
        a = i
        b = i+1
        try:
            for idx, set_items in enumerate(make_nouns_set_test):
            #print set_items
                for split_items in src_split_arr_nouns[i]:
                    #print split_items
                    if set_items == split_items:
                        U[a:b, idx] = 1
                    else :
                        pass
                    #print "no"
                    #print set_items
                    #print split_items
        except:
            traceback.print_exc()
            pass
    dump_matrix_V(U, src_split_arr_nouns, make_nouns_set_test)

def split_arr_nouns():
    split_nouns = {}
    for i in range(0, len(excel_nouns_dict)):
        if excel_nouns_dict[i] is None:
            continue
        else:
            split_val = excel_nouns_dict[i].split(",")
            split_nouns[i] = split_val
            print split_nouns
    return split_nouns
         

def get_excel_nouns():
    #wb=load_workbook('./file/reference.xlsx')
    wb=load_workbook(REFERENCE_EXCEL)
    sheetList = wb.get_sheet_names()
    #sheet = wb.get_sheet_by_name('extraction')
    sheet = wb.get_sheet_by_name(EXTRACTION_SHEET)
    row_count = sheet.get_highest_row()

    all_cellValue=[]

    for i in range(2,row_count):
        if sheet.row_dimensions[i].visible :
            cellValue = sheet.cell(row=i, column=3).value
            all_cellValue.append(cellValue)
        else :
            continue    
    return all_cellValue 

def article_id_set_dict():
    sheet = load_reference_excel()
    row_count = sheet.get_highest_row()

    dict_article_id_set={}
    article_items = set()

    for i in range(2,row_count):
        cell_article_id = sheet.cell(row=i, column=4).value   # 
        article_items.add(cell_article_id)

    article_items = list(article_items)

    for j in range(0, len(article_items)):
        dict_article_id_set[j] = article_items[j]    # dictionary   index : position 

    #print dict_article_id_set
    return dict_article_id_set

def n_informer_set_dict():

    sheet = load_reference_excel()
    row_count = sheet.get_highest_row()

    dict_n_informer_set={}
    n_informer_items = set()

    for i in range(2,row_count):
        cell_n_informer = sheet.cell(row=i, column=1).value   # 
        if cell_n_informer is None:
            continue 
        n_informer_items.add(cell_n_informer)

    n_informer_items = list(n_informer_items)

    for j in range(0, len(n_informer_items)):
        dict_n_informer_set[j] = n_informer_items[j] 
    return dict_n_informer_set


RUN_NA_EXTRACTION=0

if __name__=="__main__":
    if RUN_NA_EXTRACTION==0:
        print 'by pass na_extraction main ()...'
        
    else:
        excel_noun()
          # excel_noun processing
        try : 
            wb = load_workbook(REFERENCE_EXCEL)
            sheet = wb.get_sheet_by_name(EXTRACTION_SHEET)
        except :
            traceback.print_exc()
            excel_noun()
        
        # nouns.p file check
        #if os.path.isfile("./file/nouns.p"):
        if os.path.isfile(DICT_NOUNS):
            print " nouns.p file existed" 
            excel_nouns_dict = nt.loadObjectBinaryFast(DICT_NOUNS)
        else:
            try :
                excel_nouns = get_excel_nouns()  
                pickle.dump( excel_nouns, open( DICT_NOUNS, "wb" ) )
                #nt.saveObjectBinaryFast(excel_nouns, DICT_ORG_SET )
                print " now nouns.p file create " 
            except :
                print " nouns.p file make error " 
    
        # id, org, pos dictionary file check
        if os.path.isfile(DICT_ID_NAME) and os.path.isfile(DICT_ORG) \
        and os.path.isfile(DICT_TYPE) and os.path.isfile(DICT_POS) \
        and os.path.isfile(DICT_CODE)  and os.path.isfile(DICT_CLASSIFIED): 
            print  'Found a dictionary for news sources'
    
            # organization set dict
            src_org_set=nt.loadObjectBinaryFast(DICT_ORG_SET)
            # position set dict
            src_pos_set=nt.loadObjectBinaryFast(DICT_POS_SET)
    
            
    
            src_name=nt.loadObjectBinaryFast(DICT_ID_NAME)
            src_org=nt.loadObjectBinaryFast(DICT_ORG)
    
           # dictionary of type
            src_type=nt.loadObjectBinaryFast(DICT_TYPE)
            # dictionary of position 
            src_pos=nt.loadObjectBinaryFast(DICT_POS)
            # dictionary of code
            src_code=nt.loadObjectBinaryFast(DICT_CODE)
            """
            is_classified_paper_category
          | Y | 본 정보원이 나온 신문지면의 분류에 의해 코딩 |
          | N | 본 정보원이 직함이나, 소속에 의해 코딩이 된 것 |
            """
            src_classified=nt.loadObjectBinaryFast(DICT_CLASSIFIED) 
            
        else:
            try :
                print  'Save a dictionary for news sources'
    
                # organization set()
                org_set = org_set_dict()
                nt.saveObjectBinaryFast(org_set, DICT_ORG_SET )
    
                # position set()
    
                pos_set = pos_set_dict()
                nt.saveObjectBinaryFast(pos_set, DICT_POS_SET )
    
                excel_id_name, excel_org, excel_type, excel_pos, excel_code, excel_classified \
                = get_excel_informers()
                nt.saveObjectBinaryFast(excel_id_name, DICT_ID_NAME )
                nt.saveObjectBinaryFast(excel_org, DICT_ORG )
                nt.saveObjectBinaryFast(excel_type, DICT_TYPE )
                nt.saveObjectBinaryFast(excel_pos,  DICT_POS )
                nt.saveObjectBinaryFast(excel_code, DICT_CODE)
                nt.saveObjectBinaryFast(excel_classified,  DICT_CLASSIFIED  )
            except :
                traceback.print_exc()
                print " get_excel_informers file make error "
    
    
    
        if os.path.isfile(DICT_INFORMER):
            print " Found a class  list for news sources "
            # News Source Matrix        
            src_mat=nt.loadObjectBinaryFast(DICT_INFORMER)
        else :
            try:
                informer_tmp = informer_class_dict()
                nt.saveObjectBinaryFast(informer_tmp,DICT_INFORMER) # replace with a shorter func.       
            except :
                traceback.print_exc()
                print " informer save  error"
        
        # article id in extraction sheet
        if os.path.isfile(DICT_ARTICLE_ID_SET):
            print " Found a class  list for news sources "
            # News Source Matrix        
            src_article_id_set=nt.loadObjectBinaryFast(DICT_ARTICLE_ID_SET)
        else :
            try:
                article_id_set_tmp = article_id_set_dict()
                nt.saveObjectBinaryFast(article_id_set_tmp,DICT_ARTICLE_ID_SET) # replace with a shorter func.       
            except :
                traceback.print_exc()
                print " article id set save  error"
        # informers in extraction sheet
        if os.path.isfile(DICT_N_INFORMER_SET):
            print " Found a class  list for news sources "
            # News Source Matrix        
            src_n_informer_set=nt.loadObjectBinaryFast(DICT_N_INFORMER_SET)
        else :
            try:
                n_informer_tmp = n_informer_set_dict()
                #print sorted(n_informer_tmp.keys())
                nt.saveObjectBinaryFast(n_informer_tmp,DICT_N_INFORMER_SET) # replace with a shorter func.       
            except :
                traceback.print_exc()
                print " n informer set save  error"
    
        if os.path.isfile(DICT_SPLIT_ARR_NOUNS):
            print " Found a dict_split_arr_nouns  "
            src_split_arr_nouns=nt.loadObjectBinaryFast(DICT_SPLIT_ARR_NOUNS)
        else :
            try:
                split_arr_nouns_tmp = split_arr_nouns()
                #print sorted(n_informer_tmp.keys())
                nt.saveObjectBinaryFast(split_arr_nouns_tmp,DICT_SPLIT_ARR_NOUNS) # replace with a shorter func.       
            except :
                traceback.print_exc()
                print " dict split arr nouns save  error"
        #all_ns=get_all_NS()
        #n_informer_set_dict()
        #matrix_U()
    
        #make_nouns_set()
        #split_arr_nouns()
    
        #informer_class_dict()
    
        #matrix_V()
    
        extraction_tuple()
    
